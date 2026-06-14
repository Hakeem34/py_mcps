import os
import re
import sys
import argparse
import datetime
import json
from unittest import result
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple

import zipfile
import re
from lxml import etree as ET
from oletools.olevba import VBA_Parser

import dataclasses
from pathlib import Path
from pathlib import PurePosixPath
from mcp.server.fastmcp import FastMCP

EMU_PER_PIXEL = 9525
EMU_PER_MM    = 36000
EMU_PER_INCH  = 914400

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r"   : "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr" : "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a"   : "http://schemas.openxmlformats.org/drawingml/2006/main",
    'x14':  'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main',
}

NS_CHART  = {
    'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}

NS_TC     = {'tc': 'http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments'}
NS_PERSON = {'p': 'http://schemas.microsoft.com/office/spreadsheetml/2018/person'}

HEADER_FOOTER_TAGS = ['oddHeader','oddFooter', 'evenHeader','evenFooter', 'firstHeader','firstFooter']


@dataclasses.dataclass(slots=True)
class CellInfo:
    value    : str = ""
    formula  : str = ""
    style_Id : str = ""
    type     : str = ""

@dataclasses.dataclass(slots=True)
class FillInfo:
    patternType : str = ""
    fgColor     : dict = dataclasses.field(default_factory=dict)
    bgColor     : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class FontInfo:
    name        : str = ""
    charset     : str = ""
    family      : str = ""
    color       : dict = dataclasses.field(default_factory=dict)
    size        : int = 0

@dataclasses.dataclass(slots=True)
class BorderInfo:
    left         : dict = dataclasses.field(default_factory=dict)
    right        : dict = dataclasses.field(default_factory=dict)
    top          : dict = dataclasses.field(default_factory=dict)
    bottom       : dict = dataclasses.field(default_factory=dict)
    diagonal     : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class cellXfsInfo:
    numFmt_Id : str = ""
    font_Id   : str = ""
    fill_Id   : str = ""
    border_Id : str = ""
    xf_Id     : str = ""
    alignment : dict = dataclasses.field(default_factory=dict)
    apply     : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class cellStyleInfo:
    name      : str = ""
    numFmt_Id : str = ""
    font_Id   : str = ""
    fill_Id   : str = ""
    border_Id : str = ""
    xf_Id     : str = ""
    alignment : dict = dataclasses.field(default_factory=dict)
    apply     : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class StylesInfo:
    numFmts    : dict = dataclasses.field(default_factory=dict)
    cellStyles : list = dataclasses.field(default_factory=list)
    cellXfs    : list = dataclasses.field(default_factory=list)
    fonts      : list = dataclasses.field(default_factory=list)
    borders    : list = dataclasses.field(default_factory=list)
    fills      : list = dataclasses.field(default_factory=list)
    dxfs       : list = dataclasses.field(default_factory=list)

@dataclasses.dataclass(slots=True)
class WorkSheetXML:
    rid                : str = ""
    name               : str = ""
    xml_path           : str = ""
    dimension          : str = ""
    max_row            : int = -1
    min_row            : int = -1
    max_col            : int = -1
    min_col            : int = -1
    cell_count         : int = 0
    hidden             : bool = False
    show_grid_lines    : bool = True
    merge_cells        : list = dataclasses.field(default_factory=list)
    default_col_width  : str = ""
    default_row_height : str = ""
    outlineLevelRow    : int = 0
    outlineLevelCol    : int = 0
    header_footer      : dict = dataclasses.field(default_factory=dict)
    cells              : dict = dataclasses.field(default_factory=dict)
    comments           : list = dataclasses.field(default_factory=list)
    images             : list = dataclasses.field(default_factory=list)
    charts             : list = dataclasses.field(default_factory=list)
    shapes             : list = dataclasses.field(default_factory=list)
    shared_formulas    : dict = dataclasses.field(default_factory=dict)
    hyper_links        : list = dataclasses.field(default_factory=list)
    tables             : list = dataclasses.field(default_factory=list)
    rels               : dict = dataclasses.field(default_factory=dict)
    filter             : FilterInfo = None
    page_setup         : dict = dataclasses.field(default_factory=dict)
    row_breaks         : list = dataclasses.field(default_factory=list)
    col_breaks         : list = dataclasses.field(default_factory=list)

@dataclasses.dataclass(slots=True)
class WorkBookXML:
    wb_path        : str
    work_sheets    : dict = dataclasses.field(default_factory=dict)
    vba_macros     : list = dataclasses.field(default_factory=list)
    shared_strings : list = dataclasses.field(default_factory=list)
    person_info    : dict = dataclasses.field(default_factory=dict)
    names          : dict = dataclasses.field(default_factory=dict)
    styles_info    : StylesInfo = dataclasses.field(default_factory=StylesInfo)

@dataclasses.dataclass(slots=True)
class FilterInfo:
    ref            : str = ""
    filters        : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class TableInfo:
    rid            : str = ""
    id             : str = ""
    name           : str = ""
    disp_name      : str = ""
    ref            : str = ""
    style          : dict = dataclasses.field(default_factory=dict)
    columns        : list = dataclasses.field(default_factory=list)
    filter         : FilterInfo = None
    altText        : str = ""
    altTextSummary : str = ""


@dataclasses.dataclass(slots=True)
class DefinedName:
    name            : str = ""
    formula         : str = ""
    comment         : str = ""
    local_sheet_id  : str = ""
    hidden          : str = ""
    
@dataclasses.dataclass(slots=True)
class SharedString:
    text           : str = ""
    ruby           : str = ""

@dataclasses.dataclass(slots=True)
class CommentInfo:
    author      : str
    cell        : str
    text        : str

@dataclasses.dataclass(slots=True)
class CommentXML:
    new_comment : bool = False
    xml_path    : str = ""
    comments    : list = dataclasses.field(default_factory=list)

@dataclasses.dataclass(slots=True)
class ShapeInfo:
    id       : int = 0
    type     : str = "sp"
    name     : str = ""
    descr    : str = ""
    text     : str = ""
    col      : int = 0
    row      : int = 0
    geometry : str = ""
    off_x    : int = 0
    off_y    : int = 0
    width    : int = 0
    height   : int = 0
    children : list = dataclasses.field(default_factory=list)

@dataclasses.dataclass(slots=True)
class HyperLinkInfo:
    ref      : str = ""
    location : str = ""
    display  : str = ""
    rid      : str = ""

@dataclasses.dataclass(slots=True)
class ChartInfo:
    xml_path : str = ""
    title    : str = ""
    series   : list = dataclasses.field(default_factory=list)
    axis     : dict = dataclasses.field(default_factory=dict)

@dataclasses.dataclass(slots=True)
class VBA_macro:
    module   : str = ""
    lines    : list = dataclasses.field(default_factory=list)


g_current_wb      = None
g_log_file        = None

# FastMCPのインスタンスを作成
mcp = FastMCP()

def print_log(text, file=sys.stderr):
    print(text, file=sys.stderr)
    if g_log_file:
        print(text, file=g_log_file)
        g_log_file.flush()

def print_shape_info(shape, indent):
#   print_log(f'{' '*indent*2}[{shape.id}][{shape.name}][{shape.type}] Addr:({shape.col}, {shape.row}) offset:({shape.off_x}, {shape.off_y}) size:({shape.width}, {shape.height})')
#   for child in shape.children:
#       print_shape_info(child, indent + 1)
    return

def get_app_path():
    if getattr(sys, 'frozen', False):
        # EXEとして実行されている場合
        return os.path.dirname(os.path.abspath(sys.executable))
    else:
        # 通常のPythonスクリプトとして実行されている場合
        return os.path.dirname(os.path.abspath(__file__))

def create_log_file():
    global g_log_file

    log_path = os.path.join(get_app_path(), "mcplog")
    os.makedirs(log_path, exist_ok = True)

    now = datetime.datetime.now()
    time_stamp = now.strftime('%Y%m%d_%H%M%S')
    log_path = os.path.join(log_path, "mcp_excel_" + time_stamp + ".log")
    g_log_file = open(log_path, "w", encoding="utf-8")

def get_shape_text(sp, ns):
    paragraphs = []

    # paragraph単位
    for p in sp.xpath('.//a:p', namespaces=ns):
        runs = []
        # run単位
        for r in p.xpath('./a:r | ./a:br', namespaces=ns):
            # 改行
            if r.tag.endswith('br'):
                runs.append('\n')
                continue

            # text
            texts = r.xpath('./a:t/text()', namespaces=ns)
            if texts:
                runs.append(texts[0])

        paragraphs.append(''.join(runs))

    return '\n'.join(paragraphs)

def get_comment_text(text_elem):
    # 普通コメント:
    # <text><t>AAA</t></text>
    texts = text_elem.xpath('.//main:t', namespaces=NS)
    return ''.join(t.text for t in texts if t.text)

def get_chart_text(text_elem):
    # chart value
    vals = text_elem.xpath('.//c:v', namespaces=NS_CHART)
    if vals:
        return ''.join(v.text for v in vals if v.text)

    # drawing text
    texts = text_elem.xpath('.//a:t', namespaces=NS_CHART)
    return ''.join(t.text for t in texts if t.text)

def load_persons(z : zipfile.ZipFile):
    persons = {}

    # person.xml 探索
    for name in z.namelist():
        if not name.startswith('xl/persons/'):
            continue

        if not name.endswith('.xml'):
            continue

        root = ET.fromstring(z.read(name))
        for person in root.xpath('//p:person', namespaces=NS_PERSON):
            person_id = person.get('id')
            display_name = (person.get('displayName') or '')
#           user_id = (person.get('userId') or '')
#           persons[person_id] = {'display_name': display_name, 'user_id': user_id,}
            persons[person_id] = display_name

    return persons


def get_rels_path(parts_path: str):
    """
    .xmlファイルに対する.relsファイルのパス生成
    """
    p = PurePosixPath(parts_path)
    return str(p.parent / '_rels' / f'{p.name}.rels')

def parse_shape(sp):
    """
    図形、コネクタの解析
    """
    shape = ShapeInfo()

    cNvPr = sp.find('./xdr:nvSpPr/xdr:cNvPr', namespaces=NS)
    if cNvPr is not None:
        shape.id = cNvPr.get('id')
        shape.name = cNvPr.get('name') or ''
        shape.descr = cNvPr.get('descr') or ''

        # 幾何学情報、サイズ
        shape.geometry = sp.xpath(".//a:prstGeom/@prst", namespaces=NS)
        pos = sp.xpath(".//a:xfrm/a:off", namespaces=NS)
        if pos:
            x = int(pos[0].attrib["x"])
            y = int(pos[0].attrib["y"])
            shape.off_x = int(x / EMU_PER_MM)
            shape.off_y = int(y / EMU_PER_MM)

        size = sp.xpath(".//a:xfrm/a:ext", namespaces=NS)
        if size:
            cx = int(size[0].attrib["cx"])
            cy = int(size[0].attrib["cy"])
            shape.width  = int(cx / EMU_PER_MM)
            shape.height = int(cy / EMU_PER_MM)
    else:
        cNvPr = sp.find('./xdr:nvCxnSpPr/xdr:cNvPr', namespaces=NS)
        if cNvPr is not None:
            shape.id = cNvPr.get('id')
            shape.name = cNvPr.get('name') or ''
            shape.type = "cxnsp"
#           print_log(f'[{shape.id}][{shape.name}]:コネクタ')

    shape.text = get_shape_text(sp, NS)
    return shape

def parse_group_shape(grp):
    """
    グループ化した図形の解析
    """
    shape = ShapeInfo()
    cNvPr = grp.find('./xdr:nvGrpSpPr/xdr:cNvPr', namespaces=NS)
    if cNvPr is not None:
        shape.id = cNvPr.get('id')
        shape.name = cNvPr.get('name') or ''
        shape.type = "grpsp"

#   print_log(f'[{shape.id}][{shape.name}]')
    # group内shape: descendant shapes in a group
    for sp in grp.xpath('./xdr:sp', namespaces=NS):
#       print_log(f"xdr:sp")
        shape.children.append(parse_shape(sp))

    for cnxsp in grp.xpath('./xdr:cxnSp', namespaces=NS):
        shape.children.append(parse_shape(cnxsp))

    for subgrp in grp.xpath('./xdr:grpSp', namespaces=NS):
#       print_log(f"xdr:grpSp2")
        shape.children.append(parse_group_shape(subgrp))

    return shape

def parse_table_xml(z : zipfile.ZipFile, table_xml_path : str, ws_obj : WorkSheetXML):
    """
    テーブルの解析
    """
#   print_log(f'parse_table_xml : {table_xml_path}')
    xml = ET.fromstring(z.read(table_xml_path))

    table_info = TableInfo()

    # table attribute
    table_info.id = xml.get('id')
    table_info.name = xml.get('name')
    table_info.disp_name = xml.get('displayName')
    table_info.ref = xml.get('ref')
#   print_log(f'  TABLE: ' f'id={table_info.id} ' f'name={table_info.name} ' f'ref={table_info.ref}')

    # table style
    style = xml.find('./main:tableStyleInfo', namespaces=NS)
    if style is not None:
        table_info.style = {
            'name': style.get('name'),
            'showFirstColumn': style.get('showFirstColumn'),
            'showLastColumn': style.get('showLastColumn'),
            'showRowStripes': style.get('showRowStripes'),
            'showColumnStripes': style.get('showColumnStripes'),
        }

#       print_log(f'    STYLE: ' f'{table_info.style["name"]}')

    # columns
    for col in xml.xpath('./main:tableColumns/main:tableColumn', namespaces=NS):
        col_info = {
            'id': col.get('id'),
            'name': col.get('name'),
            'totalsRowLabel': col.get('totalsRowLabel'),
            'totalsRowFunction': col.get('totalsRowFunction'),
            'calculatedColumnFormula': None,
        }

        # calculated column formula
        calc_formula = col.find('./main:calculatedColumnFormula', namespaces=NS)
        if calc_formula is not None:
            col_info['calculatedColumnFormula'] = \
                calc_formula.text

        table_info.columns.append(col_info)

#       print_log(f'    COLUMN: ' f'id={col_info["id"]} ' f'name={col_info["name"]}')
        if col_info['calculatedColumnFormula']:
            print_log(f'      FORMULA: ' f'{col_info["calculatedColumnFormula"]}')

    # auto filter
    af = xml.find('./main:autoFilter', namespaces=NS)
    if af is not None:
        af_ref = af.get('ref')
#       print_log(f'    AUTOFILTER: {af_ref}')
        table_filter = FilterInfo()
        for fc in af.findall('./main:filterColumn', namespaces=NS):
            filter_info = {
                'colId': fc.get('colId'),
                'type': None,
                'values': [],
            }

            # normal filters
            filters = fc.find('./main:filters', namespaces=NS)
            if filters is not None:
                filter_info['type'] = 'filters'
                for f in filters.findall('./main:filter', namespaces=NS):
                    val = f.get('val')
                    filter_info['values'].append(val)

            # custom filters
            custom_filters = fc.find('./main:customFilters', namespaces=NS)
            if custom_filters is not None:
                filter_info['type'] = 'customFilters'
                for cf in custom_filters.findall('./main:customFilter', namespaces=NS):
                    filter_info['values'].append({
                        'operator': cf.get('operator'),
                        'val': cf.get('val'),
                    })

            # color filter
            color_filter = fc.find('./main:colorFilter', namespaces=NS)
            if color_filter is not None:
                filter_info['type']      = 'colorFilter'
                filter_info['dxfId']     = color_filter.get('dxfId')
                filter_info['cellColor'] = color_filter.get('cellColor')

            table_info.filter = table_filter
#           print_log(f'      FILTER: ' f'col={filter_info["colId"]} ' f'type={filter_info["type"]} ' f'value={filter_info["values"]}')

    for ext_tables in xml.xpath('./main:extLst/main:ext/x14:table', namespaces=NS):
        if ext_tables is not None:
            table_info.altText = ext_tables.get('altText', '')
            table_info.altTextSummary = ext_tables.get('altTextSummary', '')
#           print_log(f'    altText: {table_info.altText} altTextSummary: {table_info.altTextSummary}')

    # worksheet objectへ保存
    ws_obj.tables.append(table_info)


def parse_drawing_xml(z : zipfile.ZipFile, draw_xml_path : str, ws_obj : WorkSheetXML):
    """
    図形描画のxml解析
    """
    xml = ET.fromstring(z.read(draw_xml_path))
    for anchor in xml.xpath("//xdr:twoCellAnchor | //xdr:oneCellAnchor | //xdr:absoluteAnchor", namespaces=NS):
#       print_log(f'parse_anchor():{anchor}')

        # normal shape
        for sp in anchor.xpath('./xdr:sp', namespaces=NS):
#           print_log(f"xdr:sp1")
            shape = parse_shape(sp)

            # セル座標
            shape.col  = int(anchor.xpath(".//xdr:from/xdr:col/text()", namespaces=NS)[0]) + 1
            shape.row  = int(anchor.xpath(".//xdr:from/xdr:row/text()", namespaces=NS)[0]) + 1
#           print_log(f'[{shape.id}][{shape.name}] Addr:({shape.col}, {shape.row})')
            print_shape_info(shape, 0)
            ws_obj.shapes.append(shape)

        for cnxsp in anchor.xpath('./xdr:cxnSp', namespaces=NS):
            shape = parse_shape(cnxsp)

            # セル座標
            shape.col  = int(anchor.xpath(".//xdr:from/xdr:col/text()", namespaces=NS)[0]) + 1
            shape.row  = int(anchor.xpath(".//xdr:from/xdr:row/text()", namespaces=NS)[0]) + 1
#           print_log(f'[{shape.id}][{shape.name}] Addr:({shape.col}, {shape.row})')
            print_shape_info(shape, 0)
            ws_obj.shapes.append(shape)


        for grp in anchor.xpath('./xdr:grpSp', namespaces=NS):
#           print_log(f"xdr:grpSp1")
            shape = parse_group_shape(grp)
            # セル座標
            shape.col  = int(anchor.xpath(".//xdr:from/xdr:col/text()", namespaces=NS)[0]) + 1
            shape.row  = int(anchor.xpath(".//xdr:from/xdr:row/text()", namespaces=NS)[0]) + 1
#           print_log(f'[{shape.id}][{shape.name}] Addr:({shape.col}, {shape.row})')
            print_shape_info(shape, 0)
            ws_obj.shapes.append(shape)
 
    rel_path = get_rels_path(draw_xml_path)
    if rel_path in z.namelist():
#       print_log(f'[rels] {draw_xml_path} -> {rel_path}')
        rel_xml = ET.fromstring(z.read(rel_path))
        for rel in rel_xml:
            if 'chart' in rel.attrib["Type"]:
#               print_log(f'chart : {rel.attrib["Target"]}')
                chart_path = rel.attrib["Target"].replace('../', 'xl/')
                parse_chart(z, chart_path, ws_obj)
            elif 'image' in rel.attrib["Type"]:
#               print_log(f'image : {rel.attrib["Target"]}')
                ws_obj.images.append(rel.attrib["Target"])
        
    return 


def parse_vba(wb_path):
    """
    VBAマクロの解析(oletoolsにほぼお任せ)
    """
    vba_macros = []
    # 3. VBAマクロの検索
    try:
        # xlsmファイルを直接指定するだけで、内部の vbaProject.bin を自動解析
        vba_parser = VBA_Parser(wb_path)
        if vba_parser.detect_vba_macros():
            # マクロコードをストリーム抽出し、1行ずつ検索
            for subfilename, stream_path, vba_filename, vba_code in vba_parser.extract_macros():
                if vba_code:
                    macro = VBA_macro(module = vba_filename, lines = vba_code)
                    vba_macros.append(macro)
        vba_parser.close()
    except Exception as e:
        # マクロが含まれない、または破損している場合はスキップ
        pass

    return vba_macros

def parse_cell(ws_obj : WorkSheetXML, cell_elem):
    """
    セルの解析
    """
    cell = CellInfo()
    cell_addr = cell_elem.get('r')      # A1 とか
    cell_type = cell_elem.get('t')      # s, inlineStr, b など
    cell_style = cell_elem.get('s')
    val = cell_elem.find('{*}v')
    cell.type = cell_type
    if val is not None:
        cell.value = val.text
#       print_log(f'  cell1={cell_addr} type={cell_type} value={cell.value}')
    else:
        cell.value = None
#       print_log(f'  cell2={cell_addr} type={cell_type}')

    # 式の中のテキストを抽出
    formula_elem = cell_elem.find('{*}f')
    if formula_elem is not None:
        formula_type = formula_elem.get('t')
        if formula_type == 'shared':
            si = formula_elem.get('si')
            if formula_elem.text:
                ws_obj.shared_formulas[si] = formula_elem.text
                cell.formula = formula_elem.text
            else:
                cell.formula = ws_obj.shared_formulas.get(si)
        else:
            cell.formula = formula_elem.text
#           print_log(f'  cell1={cell_addr} type={cell_type} value={val.text} by formula={cell.formula}')
    else:
        cell.formula = None

    cell.style_Id = cell_style
    ws_obj.cells[cell_addr] = cell

def parse_work_sheet(z : zipfile.ZipFile, wb_obj : WorkBookXML):
    """
    ワークシートのXML解析
    """
    for ws_name, ws_obj in wb_obj.work_sheets.items():
        if (ws_obj.hidden):
            print_log(f'------------------------------------------- parse_work_sheet : [Hidden]{ws_name} ({ws_obj.xml_path})-------------------------------------------')
        else:
            print_log(f'------------------------------------------- parse_work_sheet : {ws_name} ({ws_obj.xml_path})-------------------------------------------')

        # シートのxmlを確認する
        ws_xml = ET.fromstring(z.read(ws_obj.xml_path))

        merge_cells = ws_xml.find('./main:mergeCells', namespaces=NS)
        if merge_cells is not None:
            print_log(f'  merge_cells : {merge_cells}')
            for mc in merge_cells.findall('./main:mergeCell', namespaces=NS):
                print(mc.get('ref'))
                ws_obj.merge_cells.append(mc.get('ref'))

        # sheetFormatPrのチェック
        sheet_format_pr = ws_xml.find('./main:sheetFormatPr', namespaces=NS)
        if sheet_format_pr is not None:
            ws_obj.default_col_width = sheet_format_pr.get('defaultColWidth')
            ws_obj.default_row_height = sheet_format_pr.get('defaultRowHeight')
            ws_obj.outlineLevelRow = int(sheet_format_pr.get('outlineLevelRow', '0'))
            ws_obj.outlineLevelCol = int(sheet_format_pr.get('outlineLevelCol', '0'))
#           print_log(f'  defaultColWidth: {ws_obj.default_col_width} defaultRowHeight: {ws_obj.default_row_height}, outlineLevelRow: {ws_obj.outlineLevelRow} outlineLevelCol: {ws_obj.outlineLevelCol}')

        # 枠線表示の設定
        for sv in ws_xml.xpath('//main:sheetView', namespaces=NS):
            if sv.get('showGridLines') == '0':
                ws_obj.show_grid_lines = False
            else:
                ws_obj.show_grid_lines = True
#           print_log(f'  showGridLines: {ws_obj.show_grid_lines}')

        # dimensionのチェック
        for dim in ws_xml.xpath('//main:dimension', namespaces=NS):
            ref = dim.get('ref')
#           print_log(f'  dimension: {ref}')
            ws_obj.dimension = ref

        # ヘッダ / フッタのチェック
        header_footer = ws_xml.find('./main:headerFooter', namespaces=NS)
        if header_footer is not None:
            for tag in HEADER_FOOTER_TAGS:
                elem = header_footer.find(f'./main:{tag}', namespaces=NS)
                if elem is not None and elem.text:
#                   print_log(f'  {tag}: {elem.text}')
                    ws_obj.header_footer[tag] = elem.text

        # セルのチェック
        for c in ws_xml.xpath('//main:c', namespaces=NS):
            parse_cell(ws_obj, c)

        # セルのアドレスから、シートの最大行/列と有効な値を持つセル数をカウントする
        for cell_addr, cell in ws_obj.cells.items():
            col_str = re.sub(r'\d', '', cell_addr)
            row_str = re.sub(r'\D', '', cell_addr)
            col = column_index_from_string(col_str)
            row = int(row_str)
            if ws_obj.max_row == -1 or row > ws_obj.max_row:
                ws_obj.max_row = row
            if ws_obj.min_row == -1 or row < ws_obj.min_row:
                ws_obj.min_row = row
            if ws_obj.max_col == -1 or col > ws_obj.max_col:
                ws_obj.max_col = col
            if ws_obj.min_col == -1 or col < ws_obj.min_col:
                ws_obj.min_col = col

            if cell.value is not None:
                ws_obj.cell_count += 1

        #シートに紐づくdrawings / commentsをチェックする
        rel_path = get_rels_path(ws_obj.xml_path)
#       print_log(f"check for {rel_path} from {ws_obj.xml_path}")
        if rel_path in z.namelist():
            rel_xml = ET.fromstring(z.read(rel_path))
            for rel in rel_xml:
                ws_obj.rels[rel.attrib["Id"]] = rel.attrib["Target"]
                if "drawing" in rel.attrib["Type"]:
                    draw_xml_path = "xl/drawings/" + rel.attrib["Target"].split("/")[-1]
#                   print_log(rel.attrib["Target"])
#                   print_log(f'[rels] {ws_name} -> {draw_xml_path}', file=sys.stderr)
                    parse_drawing_xml(z, draw_xml_path, ws_obj)
                elif "threadedComment" in rel.attrib["Type"]:
#                   print_log(rel.attrib["Target"])
                    comment_obj = CommentXML(new_comment=True)
                    comment_obj.xml_path = "xl/threadedComments/" + rel.attrib["Target"].split("/")[-1]
                    parse_comments(z, wb_obj, comment_obj)
                    ws_obj.comments.append(comment_obj)
                elif "comment" in rel.attrib["Type"]:
#                   print_log(rel.attrib["Target"])
                    comment_obj = CommentXML(new_comment=False)
                    comment_obj.xml_path = "xl/" + rel.attrib["Target"].split("/")[-1]
                    parse_comments(z, wb_obj, comment_obj)
                    ws_obj.comments.append(comment_obj)
                elif "table" in rel.attrib["Type"]:
                    table_xml_path = "xl/tables/" + rel.attrib["Target"].split("/")[-1]
                    parse_table_xml(z, table_xml_path, ws_obj)
                    pass
#               else:
#                   print_log(f'Other Rel[{rel.attrib["Id"]}][{rel.attrib["Type"]}] : {rel.attrib["Target"]}')

        # ハイパーリンクのチェック
        for hl in ws_xml.xpath('//main:hyperlink', namespaces=NS):
            hl_info = HyperLinkInfo()
            hl_info.ref = hl.get('ref')
            hl_info.display = hl.get('display')
            hl_info.location = hl.get('location')
            hl_info.rid = hl.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
#           if hl_info.display is not None:
#               print_log(f'HyperLink[{hl_info.ref}]:{hl_info.display}')
#           else:
#               print_log(f'HyperLink[{hl_info.ref}]:{hl_info.rid} -> {ws_obj.rels[hl_info.rid]}')

            ws_obj.hyper_links.append(hl_info)
        
        # オートフィルタの情報取得
        af = ws_xml.find('{*}autoFilter')
        if af is not None:
            af_info = FilterInfo()
            af_info.ref = af.get('ref')
#           print_log(f'AUTOFILTER RANGE: {af_info.ref}')
            for fc in af.findall('{*}filterColumn'):
                col_id = fc.get('colId')
#               print_log(f'  COLUMN: {col_id}')
                af_info.filters[col_id] = []

                # とりあえず値の選択フィルタのみParseする。他にもカラーフィルタやカスタムフィルタとかあるけど、今は無視
                for f in fc.findall('.//{*}filter'):
                    val = f.get('val')
#                   print_log(f'    FILTER: {val}')
                    af_info.filters[col_id].append(val)
            ws_obj.filter = af_info

        # page setup / print settings
        page_setup_elem = ws_xml.find('{*}pageSetup')
        if page_setup_elem is not None:
            for key in ('scale', 'fitToPage', 'fitToWidth', 'fitToHeight', 'orientation', 'paperSize'):
                value = page_setup_elem.get(key)
                if value is not None:
                    ws_obj.page_setup[key] = value

        for brk in ws_xml.findall('.//{*}rowBreaks/{*}brk'):
            row_id = brk.get('id')
            if row_id is not None:
                try:
                    ws_obj.row_breaks.append(int(row_id))
                except Exception:
                    pass

        for brk in ws_xml.findall('.//{*}colBreaks/{*}brk'):
            col_id = brk.get('id')
            if col_id is not None:
                try:
                    ws_obj.col_breaks.append(int(col_id))
                except Exception:
                    pass
    return


def parse_chart(z : zipfile.ZipFile, chart_path : str, ws_obj : WorkSheetXML):
    """
    チャート(グラフ)のxml解析
    """
    chart_xml = ET.fromstring(z.read(chart_path))

    chart_info = ChartInfo(xml_path = chart_path)
    # chart title
    title_elem = chart_xml.find('.//c:title', namespaces=NS_CHART)
    if title_elem is not None:
        title = get_chart_text(title_elem)
#       print_log(f'title = {title}')
        chart_info.title = title

    # all series
    for ser in chart_xml.xpath('//c:ser', namespaces=NS_CHART):
        # series name / 系列名
        tx = ser.find('.//c:tx', namespaces=NS_CHART)
        if tx is not None:
            text = get_chart_text(tx)
            if text:
#               print_log(f'series name = {text}')
                chart_info.series.append(text)
        
    # category axis title / カテゴリ軸ラベル
    for ax in chart_xml.xpath('//c:catAx', namespaces=NS_CHART):
        texts = ax.xpath('.//c:title//a:t', namespaces=NS_CHART)
        title = ''.join(t.text for t in texts if t.text)
        if title:
#           print_log(f'cat_axis_title = {title}')
            chart_info.axis['category'] = title

    # value axis title / 値の軸ラベル
    for ax in chart_xml.xpath('//c:valAx', namespaces=NS_CHART):
        texts = ax.xpath('.//c:title//a:t', namespaces=NS_CHART)
        title = ''.join(t.text for t in texts if t.text)
        if title:
#           print_log(f'val_axis_title = {title}')
            chart_info.axis['value'] = title

    ws_obj.charts.append(chart_info)
    return chart_info
 
def parse_styles(z : zipfile.ZipFile, wb_obj : WorkBookXML):
    """
    スタイルのxml解析
    """
    print_log(f'------------------------------------------- parse_styles -------------------------------------------')
    if 'xl/styles.xml' in z.namelist():
        styles_xml = ET.fromstring(z.read('xl/styles.xml'))
        # number formats
        for num_fmt in styles_xml.xpath('//main:numFmts/main:numFmt', namespaces=NS):
            numFmtId = num_fmt.get('numFmtId')
            formatCode = num_fmt.get('formatCode')
            wb_obj.styles_info.numFmts[numFmtId] = formatCode.replace('""', '')    # ”コ””メ””ン””ト”などの単文字を連結してテキストに戻す
            print_log(f'numFmtId={numFmtId} formatCode={wb_obj.styles_info.numFmts[numFmtId]}')
        
        # fonts
        for font in styles_xml.xpath('//main:fonts/main:font', namespaces=NS):
            font_info = FontInfo()
            name = font.find('./main:name', namespaces=NS)
            if name is not None:
                font_info.name = name.get('val')
            color = font.find('./main:color', namespaces=NS)
            if color is not None:
                font_info.color = color.attrib
            sz = font.find('./main:sz', namespaces=NS)
            if sz is not None:
                font_info.size = sz.get('val')
            charset = font.find('./main:charset', namespaces=NS)
            if charset is not None:
                font_info.charset = charset.get('val')
            family = font.find('./main:family', namespaces=NS)
            if family is not None:
                font_info.family = family.get('val')

            print_log(f'font name={font_info.name}, color={font_info.color}, size={font_info.size}, charset={font_info.charset}, family={font_info.family}')
            wb_obj.styles_info.fonts.append(font_info)
        
        # fills
        for fill in styles_xml.xpath('//main:fills/main:fill', namespaces=NS):
            fill_info = FillInfo()
            pattern_fill = fill.find('./main:patternFill', namespaces=NS)
            if pattern_fill is not None:
                fill_info.patternType = pattern_fill.get('patternType')
                fg_color = pattern_fill.find('./main:fgColor', namespaces=NS)
                if fg_color is not None:
                    fill_info.fgColor = fg_color.attrib
                bg_color = pattern_fill.find('./main:bgColor', namespaces=NS)
                if bg_color is not None:
                    fill_info.bgColor = bg_color.attrib
            print_log(f'fill patternType={fill_info.patternType}, fg_color={fill_info.fgColor}, bg_color={fill_info.bgColor}')
            wb_obj.styles_info.fills.append(fill_info)

        # borders
        for border in styles_xml.xpath('//main:borders/main:border', namespaces=NS):
            border_info = BorderInfo()
            for position in ['left', 'right', 'top', 'bottom']:
                border_side = border.find(f'./main:{position}', namespaces=NS)
                if border_side is not None:
                    border_info.__getattribute__(position)['style'] = border_side.attrib.get('style')
                    color = border.find(f'./main:{position}/main:color', namespaces=NS)
                if color is not None:
#                   print_log(f'border left color={color.attrib}')
                    border_info.__getattribute__(position)['color'] = color.attrib
            if border.attrib.get('diagonalUp') == '1':
                border_info.diagonal["up"] = True
            else:
                border_info.diagonal["up"] = False

            if border.attrib.get('diagonalDown') == '1':
                border_info.diagonal["down"] = True
            else:
                border_info.diagonal["down"] = False

#           print_log(f'border diagonal up={border_info.diagonal["up"]}, down={border_info.diagonal["down"]}, l={border_info.left}, r={border_info.right}, t={border_info.top}, b={border_info.bottom}')
            wb_obj.styles_info.borders.append(border_info)

        # cellStyleXfs
        for cell_style_xf in styles_xml.xpath('//main:cellStyleXfs/main:xf', namespaces=NS):
            cell_style_info = cellStyleInfo()
            cell_style_info.numFmt_Id = cell_style_xf.get('numFmtId')

#           print_log(f'cellStyleXfs numFmtId={cell_style_info.numFmt_Id}')
            wb_obj.styles_info.cellStyles.append(cell_style_info)


def parse_shared_string(z : zipfile.ZipFile, wb_obj : WorkBookXML):
    """
    共有文字列のxml解析
    """
    shared_strings = []
    print_log(f'------------------------------------------- parse_shared_string -------------------------------------------')
    if 'xl/sharedStrings.xml' in z.namelist():
        string_index = 0
        ss_xml = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for si in ss_xml.xpath('//main:si', namespaces=NS):
            full_text = ''.join(t.text for t in si.xpath('.//main:t[not(ancestor::main:rPh)]', namespaces=NS)if t.text)
            ruby_text = ''.join(t.text for t in si.xpath('.//main:t[ancestor::main:rPh]', namespaces=NS)if t.text)
#           print_log(full_text)
            ss = SharedString()
            ss.text = full_text
            ss.ruby = ruby_text
            shared_strings.append(ss)
    
    wb_obj.shared_strings = shared_strings
    return

def parse_comments(z : zipfile.ZipFile, wb_obj : WorkBookXML, comment_obj : CommentXML):
    """
    コメント・メモのxml解析
    """
    comments = []
    print_log(f'------------------------------------------- parse_comments {comment_obj.xml_path}-------------------------------------------')
    comment_xml = ET.fromstring(z.read(comment_obj.xml_path))
    if comment_obj.new_comment:
        for tc in comment_xml.xpath('//tc:threadedComment', namespaces=NS_TC):
            # セル位置
            cell_ref = tc.get('ref', '')
#           comment_id = tc.get('id', '')
#           parent_id = tc.get('parentId', '')
            person_id = tc.get('personId', '')
            author = ''
            if person_id in wb_obj.person_info:
                author = wb_obj.person_info[person_id]
#           dt = tc.get('dT', '')
            text_elem = tc.find('{*}text')
            text = ''
            if text_elem is not None and text_elem.text:
                text = ''.join(text_elem.itertext())

#           print_log(text)
            comment_obj.comments.append(CommentInfo(author, cell_ref, text))
    else:
        authors = []
        for author in comment_xml.xpath('//main:authors/main:author', namespaces=NS):
            authors.append(author.text or '')

        for comment in comment_xml.xpath('//main:comment', namespaces=NS):
            text_elem = comment.find('{*}text')
            cell_ref = comment.get('ref')
            text = get_comment_text(text_elem)
            author_name = ""
            author_id = comment.get('authorId')
            if author_id is not None:
                idx = int(author_id)
                if 0 <= idx < len(authors):
                    author_name = authors[idx]

            comment_obj.comments.append(CommentInfo(author_name, cell_ref, text))
#           print_log(f'[{cell_ref}] : {text}')
    return

def parse_work_book(wb_path : str, z : zipfile.ZipFile):
    """
    ワークブックのxml解析
    """
    print_log(f'------------------------------------------- parse_work_book {wb_path}-------------------------------------------')

    # workbook.xmlを読む
    wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
    # workbook.xml.relsを読む
    wb_rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))

    # rId → sheetX.xml のマップを作る
    rid_to_sheetxml = {}
    for rel in wb_rels:
        if "worksheet" in rel.attrib["Type"] or "chartsheet" in rel.attrib["Type"]:
            rid_to_sheetxml[rel.attrib["Id"]] = rel.attrib["Target"]
#   print_log(rid_to_sheetxml)

    wb_obj = WorkBookXML(wb_path = wb_path)
    wb_obj.person_info = load_persons(z)
    for sheet in wb_xml.xpath("//main:sheet", namespaces=NS):
        rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        if rid in rid_to_sheetxml:
            ws_obj = WorkSheetXML()
            ws_obj.name = sheet.attrib["name"]
            ws_obj.rid = rid
            ws_obj.xml_path = "xl/" + rid_to_sheetxml[ws_obj.rid]
            if "state" in sheet.attrib:
                ws_obj.hidden = (sheet.attrib["state"] == "hidden")
            wb_obj.work_sheets[ws_obj.name] = ws_obj
        else:
            print_log(f'Unknown Sheet Type : rid={ws_obj.rid}')

    # 名前の情報を種痘
    for def_name in wb_xml.xpath("//main:definedName ", namespaces=NS):
        name_info = DefinedName()
        name_info.name    = def_name.get('name') 
        name_info.formula = def_name.text
        name_info.comment = def_name.get('comment')
        name_info.hidden = def_name.get('hidden')
        name_info.local_sheet_id = def_name.get('localSheetId')
#       print_log(f'{name_info.name}:{name_info.formula}')
        name_key = name_info.name
        if name_info.local_sheet_id:
            name_key = f"{name_info.name}:{name_info.local_sheet_id}"
        wb_obj.names[name_key] = name_info


    parse_shared_string(z, wb_obj)
    parse_styles(z, wb_obj)
    parse_work_sheet(z, wb_obj)
    return wb_obj

def parse_by_xml(wb_path):
    """
    xmlによるExcelファイル解析
    """
    with zipfile.ZipFile(wb_path) as z:
        wb_obj            = parse_work_book(wb_path, z)
        wb_obj.vba_macros = parse_vba(wb_path)
    return wb_obj


def convert_all_options():
    return "BNSVFDGCTMHUP"

def _cell_position(address : str):
    m = re.match(r'^([A-Z]+)(\d+)$', address)
    if not m:
        return None, None
    col_text, row_text = m.groups()
    col = column_index_from_string(col_text)  # これで正しいかどうかもチェック
    return col, int(row_text)

def _cell_address(col : int, row : int):
    return f"{get_column_letter(col)}{row}"

def _parse_range_ref(ref : str):
    if not ref:
        return None
    ref = re.sub(r"^([^!]+!)", "", ref)
    ref = ref.replace('$', '')
    parts = ref.split(':')
    if len(parts) != 2:
        return None
    start = parts[0]
    end = parts[1]
    start_col, start_row = _cell_position(start)
    end_col, end_row = _cell_position(end)
    if start_col is None or end_col is None:
        return None
    return start_col, start_row, end_col, end_row

def _sheet_local_id(wb_obj : WorkBookXML, sheet_name : str):
    for idx, name in enumerate(wb_obj.work_sheets):
        if name == sheet_name:
            return str(idx)
    return ""

def _get_print_area(wb_obj : WorkBookXML, sheet_name : str) -> str:
    sheet_id = _sheet_local_id(wb_obj, sheet_name)
    areas = []
    for name_info in wb_obj.names.values():
        if name_info.name == 'Print_Area' and (name_info.local_sheet_id == sheet_id or name_info.local_sheet_id == ""):
            if name_info.formula:
                areas.append(name_info.formula)
    return ';'.join(areas)

def _get_print_page_count(ws_obj : WorkSheetXML) -> int:
    if ws_obj.row_breaks or ws_obj.col_breaks:
        return max(1, len(ws_obj.row_breaks) + 1) * max(1, len(ws_obj.col_breaks) + 1)
    return 1 if ws_obj.cells else 0


def _get_cell_range(ws_obj : WorkSheetXML):
    """
    有効なセルの範囲をA1形式で返す。セルがない場合はNoneを返す。
    """
    if ws_obj.min_col >= 0 and ws_obj.min_row >= 0 and ws_obj.max_col >= 0 and ws_obj.max_row >= 0:
        range = get_column_letter(ws_obj.min_col) + str(ws_obj.min_row) + ":" + get_column_letter(ws_obj.max_col) + str(ws_obj.max_row)
    else:
        range = None

    return range

def _get_first_cell_value(ws_obj : WorkSheetXML, wb_obj : WorkBookXML):
    cell_range = _get_cell_range(ws_obj)
    if not cell_range:
        return ""
    parsed = _parse_range_ref(cell_range)
    if not parsed:
        return ""
    start_col, start_row, _, _ = parsed
    first_addr = _cell_address(start_col, start_row)
    return _cell_display_text(ws_obj.cells.get(first_addr), wb_obj)

def _get_autofilter_headers(ws_obj : WorkSheetXML, wb_obj : WorkBookXML):
    if not ws_obj.filter or not ws_obj.filter.ref:
        return []
    parsed = _parse_range_ref(ws_obj.filter.ref)
    if not parsed:
        return []
    start_col, start_row, end_col, _ = parsed
    headers = []
    for col in range(start_col, end_col + 1):
        addr = _cell_address(col, start_row)
        headers.append(_cell_display_text(ws_obj.cells.get(addr), wb_obj))
    return headers

def _compile_search_pattern(key_word : str, regex : bool):
    try:
        if regex:
            return re.compile(key_word, re.IGNORECASE)
        return re.compile(re.escape(key_word), re.IGNORECASE)
    except re.error:
        return None

def _cell_display_text(cell : CellInfo, wb_obj : WorkBookXML) -> str:
    if cell is None:
        return ""
    value = cell.value or ""
    if cell.type == 's' and value:
        try:
            idx = int(value)
            if 0 <= idx < len(wb_obj.shared_strings):
                return wb_obj.shared_strings[idx].text
        except Exception:
            pass
    return value

def _range_to_addr_list(range_addr: str) -> list:
    ranges = range_addr.split(':')
    top_left = ranges[0]
    bottom_right = ranges[1] if len(ranges) > 1 else top_left
    min_row, min_col = coordinate_to_tuple(top_left)
    max_row, max_col = coordinate_to_tuple(bottom_right)

    addr_list = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
#           addr_list.append(f"{get_column_letter(col)}{row}")
            addr_list.append(_cell_address(col, row))

    return addr_list


def _matches(pattern, text : str) -> bool:
    if not text:
        return False
    return bool(pattern.search(text))

@mcp.tool()
def grep_work_book(target_path : str, key_word : str, regex : bool = False, option : str = "all") -> str:
    """
    Search for a keyword in an Excel file.
     - target_path: Target file path.
     - key_word: Keyword to search for. If regex is true, treat this as a regular expression.
     - regex: Whether to treat the key_word as a regular expression.
     - option: Search target. "all" searches in all content, or you can specify a specific target using the following codes:
        N:Defined Name
        S:Sheet Name
        V:Cell Value
        F:Cell Formula
        D:Shape or drawing
        G:Graph, Chart
        C:Comment
        T:Table
        M:VBA, Macro
        H:HyperLink
        U:Custom number format
        P:Header / Footer
        Example: If you want to search for the keyword only in sheet names and cell values, specify "SV".
    Return value: A string containing the search results. Each line contains the file path, sheet name (if applicable), cell address (if applicable), and the content where the keyword was found.
    Example return value:
        [V]sample/test.xlsx:Sheet1!A1:This is a sample text.
        [F]sample/test.xlsx:Sheet2!B2:=SUM(A1:A10)
        [C]sample/test.xlsx:Sheet3!Comment1:This is a comment.
    Note: The actual format of the return value can be designed as needed, but it should contain enough information to identify where the keyword was found.
    """
    print_log(f'grep_work_book: target_path={target_path}, key_word={key_word}, regex={regex}, option={option}')
    if option == "all":
        option = convert_all_options()

    option_set = set(option.upper())
    matcher = _compile_search_pattern(key_word, regex)
    if matcher is None:
        return ""

    results = []
    try:
        wb_obj = parse_by_xml(target_path)
    except Exception as e:
        print_log(f'grep_work_book parse error: {e}')
        return ""

    if 'B' in option_set:
        file_name = os.path.basename(target_path)
        if _matches(matcher, file_name) or _matches(matcher, target_path):
            results.append(f"[B]{target_path}:{file_name}")

    if 'U' in option_set:
        for id, value in wb_obj.styles_info.numFmts.items():
            print_log(f'num_fmt({id}): {value}')
            if _matches(matcher, value):
                results.append(f"[U]{target_path}:{value}")

    if 'N' in option_set:
        for name_info in wb_obj.names.values():
            if _matches(matcher, name_info.name):
                results.append(f"[N]{target_path}:{name_info.name}:Name={name_info.name}")
            if _matches(matcher, name_info.formula or ""):
                results.append(f"[N]{target_path}:{name_info.name}:Formula={name_info.formula}")
            if _matches(matcher, name_info.comment or ""):
                results.append(f"[N]{target_path}:{name_info.name}:Comment={name_info.comment}")

    for ws_name, ws_obj in wb_obj.work_sheets.items():
        if 'S' in option_set and _matches(matcher, ws_name):
            results.append(f"[S]{target_path}:{ws_name}:{ws_name}")

        if 'V' in option_set or 'F' in option_set:
            for cell_addr, cell in ws_obj.cells.items():
                if 'V' in option_set:
                    text = _cell_display_text(cell, wb_obj)
                    if _matches(matcher, text):
                        results.append(f"[V]{target_path}:{ws_name}!{cell_addr}:{text}")
                if 'F' in option_set and cell.formula:
                    if _matches(matcher, cell.formula):
                        results.append(f"[F]{target_path}:{ws_name}!{cell_addr}:{cell.formula}")

        if 'D' in option_set:
            for shape in ws_obj.shapes:
                if _matches(matcher, shape.name or "") or _matches(matcher, shape.text or "") or _matches(matcher, shape.type or "") or _matches(matcher, shape.descr or ""):
                    ident = shape.name or str(shape.id)
                    content = f"{shape.type}:{shape.text or ''}".strip()
                    results.append(f"[D]{target_path}:{ws_name}!{ident}:{content}")

        if 'G' in option_set:
            for chart in ws_obj.charts:
                matched = False
                chart_texts = [chart.title] + chart.series + list(chart.axis.values())
                for text in chart_texts:
                    if _matches(matcher, text or ""):
                        matched = True
                        break
                if matched:
                    display = chart.title or chart.xml_path
                    results.append(f"[G]{target_path}:{ws_name}!{display}:{chart.xml_path}")

        if 'C' in option_set:
            for comment_obj in ws_obj.comments:
                for comment in comment_obj.comments:
                    if _matches(matcher, comment.text or "") or _matches(matcher, comment.author or "") or _matches(matcher, comment.cell or ""):
                        author_part = comment.author or ""
                        results.append(f"[C]{target_path}:{ws_name}!{comment.cell}:{author_part}:{comment.text}")

        if 'T' in option_set:
            for table_info in ws_obj.tables:
                found = False
                fields = [table_info.name, table_info.disp_name, table_info.ref, table_info.altText, table_info.altTextSummary]
                if table_info.style:
                    fields.extend([str(table_info.style.get('name', '')), str(table_info.style.get('showFirstColumn', '')), str(table_info.style.get('showLastColumn', '')), str(table_info.style.get('showRowStripes', '')), str(table_info.style.get('showColumnStripes', ''))])
                for col in table_info.columns:
                    fields.append(col.get('name', ''))
                    if col.get('calculatedColumnFormula'):
                        fields.append(col.get('calculatedColumnFormula'))
                for value in fields:
                    if _matches(matcher, value or ""):
                        found = True
                        break
                if found:
                    name = table_info.name or table_info.disp_name or table_info.ref or "Table"
                    results.append(f"[T]{target_path}:{ws_name}!{name}:{table_info.ref}")

        if 'H' in option_set:
            for hl in ws_obj.hyper_links:
                if _matches(matcher, hl.ref or "") or _matches(matcher, hl.display or "") or _matches(matcher, hl.location or ""):
                    ident = hl.ref or hl.location or hl.display or "HyperLink"
                    display = hl.display or hl.location or ""
                    results.append(f"[H]{target_path}:{ws_name}!{ident}:{display}")
        
        if 'P' in option_set:
            header_footer = ws_obj.header_footer
            for tag in HEADER_FOOTER_TAGS:
#               print_log(f'header_footer.{tag}: {header_footer.get(tag, "")}')
                if tag in header_footer and _matches(matcher, header_footer[tag] or ""):
                    results.append(f"[P]{target_path}:{ws_name}:{tag}:{header_footer[tag]}")

    if 'M' in option_set:
        for macro in wb_obj.vba_macros:
            if _matches(matcher, macro.module or ""):
                results.append(f"[M]{target_path}:{macro.module}:module")
            lines = macro.lines
            if isinstance(lines, str):
                lines = lines.splitlines()
            for line in lines or []:
                if _matches(matcher, line or ""):
                    results.append(f"[M]{target_path}:{macro.module}:{line.strip()}")

    return '\n'.join(results)


@mcp.tool()
def grep_work_books(target_path : str, key_word : str, recursive : bool = True, regex : bool = False, option : str = "all") -> str:
    """
    Search for a keyword in Excel files under the target path.
     - target_path: Target file or directory path. If a directory is specified, search for Excel files in the directory. If recursive is true, search for Excel files in subdirectories as well.
     - key_word: Keyword to search for. If regex is true, treat this as a regular expression.
     - recursive: Whether to search for Excel files in subdirectories when a directory is specified as the target path.
     - regex: Whether to treat the key_word as a regular expression.
     - option: Search target. "all" searches in all content, or you can specify a specific target using the following codes:
        B:Book File Name
        N:Defined Name
        S:Sheet Name
        V:Cell Value
        F:Cell Formula
        D:Shape or drawing
        G:Graph, Chart
        C:Comment
        T:Table
        M:VBA, Macro
        H:HyperLink
        U:Custom number format
        P:Header / Footer
        Example: If you want to search for the keyword only in sheet names and cell values, specify "SV".
     Return value: A string containing the search results. Each line contains the file path, sheet name (if applicable), cell address (if applicable), and the content where the keyword was found.
     Example return value:
        [V]sample/test.xlsx:Sheet1!A1:This is a sample text.
        [F]sample/test.xlsx:Sheet2!B2:=SUM(A1:A10)
        [C]sample/test.xlsx:Sheet3!Comment1:This is a comment.
     Note: The actual format of the return value can be designed as needed, but it should contain enough information to identify where the keyword was found.
    """

    print_log(f'grep_work_books: target_path={target_path}, key_word={key_word}, recursive={recursive}, regex={regex}, option={option}')
    grep_results = []
    if os.path.isdir(target_path):
        for root, dirs, files in os.walk(target_path):
            if not recursive:
                dirs.clear()
            for file_name in files:
                if file_name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm', '.xlam')):
                    file_path = os.path.join(root, file_name)
                    result = grep_work_book(file_path, key_word, regex, option)
                    if result:
                        grep_results.append(result)
    else:
        result = grep_work_book(target_path, key_word, regex, option)
        if result:
            grep_results.append(result)

    return '\n'.join(grep_results)

@mcp.tool()
def get_work_sheet_summary(wb_path : str, sheet_name : str = "") -> str:
    """
    Get a summary of a worksheet or all worksheets in the workbook.
     - wb_path: Workbook file path.
     - sheet_name: Optional worksheet name. If omitted, summaries for all sheets are returned.
    """
    global g_current_wb

    print_log(f'get_work_sheet_summary: wb_path={wb_path}, sheet_name={sheet_name}')
    if g_current_wb is None or g_current_wb.wb_path != wb_path:
        g_current_wb = parse_by_xml(wb_path)

    wb_obj = g_current_wb
    targets = []
    if sheet_name:
        if sheet_name not in wb_obj.work_sheets:
            return f"Sheet not found: {sheet_name}"
        targets = [sheet_name]
    else:
        targets = list(wb_obj.work_sheets.keys())

    summaries = []
    for ws_name in targets:
        ws_obj = wb_obj.work_sheets[ws_name]
        summary = []
        summary.append(f"Sheet name: {ws_name}")
        if ws_obj.hidden:
            summary.append("State: hidden")

        heading = _get_first_cell_value(ws_obj, wb_obj)
        summary.append(f"Header cell: {heading if heading else 'None'}")

        valid_range = _get_cell_range(ws_obj)
        summary.append(f"Used cell range: {valid_range if valid_range else 'None'}")

        shape_count = len(ws_obj.shapes)
        if shape_count:
            names = [shape.name for shape in ws_obj.shapes if shape.name]
            summary.append(f"Shapes: {shape_count} ({', '.join(names)})")
        else:
            summary.append("Shapes: 0")

        table_count = len(ws_obj.tables)
        if table_count:
            titles = [table_info.disp_name or table_info.name or table_info.ref or "Table" for table_info in ws_obj.tables]
            summary.append(f"Tables: {table_count} ({', '.join(titles)})")
        else:
            summary.append("Tables: 0")

        chart_count = len(ws_obj.charts)
        if chart_count:
            titles = [chart.title or chart.xml_path for chart in ws_obj.charts]
            summary.append(f"Charts: {chart_count} ({', '.join(titles)})")
        else:
            summary.append("Charts: 0")

        if ws_obj.filter and ws_obj.filter.ref:
            headers = _get_autofilter_headers(ws_obj, wb_obj)
            header_label = ', '.join(headers) if any(headers) else 'None'
            summary.append(f"AutoFilter: Yes ({ws_obj.filter.ref}) Headers: {header_label}")

        if ws_obj.outlineLevelRow != 0 or ws_obj.outlineLevelCol != 0:
            summary.append(f"Outline levels - Row: {ws_obj.outlineLevelRow}, Column: {ws_obj.outlineLevelCol}")

        if ws_obj.comments:
            comment_summaries = []
            for comment_obj in ws_obj.comments:
                for comment in comment_obj.comments:
                    author_part = f"{comment.author}:" if comment.author else ""
                    comment_summaries.append(f"{author_part}{comment.cell}:{comment.text}")
            summary.append(f"Comments: {len(ws_obj.comments)} ({'; '.join(comment_summaries)})")
        
        if ws_obj.merge_cells:
            summary.append(f"Merged cells: {len(ws_obj.merge_cells)} ({', '.join(ws_obj.merge_cells)})")

#       print_area = _get_print_area(wb_obj, ws_name)
#       summary.append(f"Print area: {'Yes (' + print_area + ')' if print_area else 'No'}")

#       summary.append(f"Print page count: {_get_print_page_count(ws_obj)}")
#       scale = ws_obj.page_setup.get('scale') or ''
#       summary.append(f"Print scale: {scale if scale else 'None'}")

        summaries.append('\n'.join(summary))

    return '\n\n'.join(summaries)

@mcp.tool()
def get_work_sheet_list(wb_path : str) -> str:
    """
    Get the list of sheets in the workbook.
    """
    global g_current_wb

    print_log(f'get_work_sheet_list: wb_path={wb_path}')
    if g_current_wb is None or g_current_wb.wb_path != wb_path:
        g_current_wb = parse_by_xml(wb_path)

    result = []
    for ws in g_current_wb.work_sheets:
        result.append(ws)

    return '\n'.join(result)

@mcp.tool()
def get_cell_values(wb_path : str, sheet_name : str, cell_range : str) -> str:
    """
    Get the value of a specific cells with valid value.
     - wb_path: Workbook file path.
     - sheet_name: Worksheet name.
     - cell_range: Cell range in A1 format (e.g., "B2:D5").
    """
    global g_current_wb

    print_log(f'get_cell_values: wb_path={wb_path}, sheet_name={sheet_name}, cell_range={cell_range}')
    if g_current_wb is None or g_current_wb.wb_path != wb_path:
        g_current_wb = parse_by_xml(wb_path)

    wb_obj = g_current_wb
    if sheet_name not in wb_obj.work_sheets:
        return f"Sheet not found: {sheet_name}"
    
    addrs = _range_to_addr_list(cell_range)  # これでセル範囲の形式が正しいかどうかもチェック
    ws_obj = wb_obj.work_sheets[sheet_name]
    results = {}
    for cell_addr in addrs:
        cell = ws_obj.cells.get(cell_addr)
        if cell and cell.value is not None:
            value = _cell_display_text(cell, wb_obj)
#           print_log(f'{cell_addr}: {value}')
            results[cell_addr] = value

    json_value = json.dumps(results, ensure_ascii=False)
    print_log(f'Cell values (json): {json_value}')
    return f"Cell range not found: {cell_range}"

def main():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-v", "--verbose", action='store_true')
    args = parser.parse_args()

    if args.verbose:
        create_log_file()

#   parse_by_xml("sample\\test_macro.xlsm")
#   parse_by_xml("sample\\test_new_comment.xlsx")
#   result = get_work_sheet_list("sample\\test_macro.xlsm")
#   print_log(f'get_work_sheet_list:\n{result}')
#   result = grep_work_books("sample", "コメント", recursive=True, regex=False, option="all")
#   print_log(f'grep_work_books:\n{result}')
#   result = get_work_sheet_summary("sample\\test_macro.xlsm", sheet_name=None)
#   print_log(f'get_work_sheet_summary:\n{result}')
#   result = get_cell_values("sample\\test_macro.xlsm", "オートフィルタ", "B2:G22")
#   print_log(f'get_cell_values:\n{result}')

    mcp.run()


if __name__ == "__main__":
    main()

